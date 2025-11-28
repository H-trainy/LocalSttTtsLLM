"""
Check Processing Progress
Shows how many rows have been processed and how many remain
"""

import os
from openpyxl import load_workbook


def check_progress(source_file='Transcript-24-11-2025.xlsx', output_file='IntentOfthetranscribetext.xlsx'):
    """Check processing progress"""
    
    print("="*60)
    print("PROCESSING PROGRESS CHECK")
    print("="*60)
    
    # Check source file
    if not os.path.exists(source_file):
        print(f"❌ Source file not found: {source_file}")
        return
    
    wb_source = load_workbook(source_file, read_only=True)
    ws_source = wb_source.active
    total_rows = ws_source.max_row - 1  # Exclude header
    wb_source.close()
    
    print(f"\n📖 Source File: {source_file}")
    print(f"   Total rows: {total_rows}")
    
    # Check output file
    if not os.path.exists(output_file):
        print(f"\n💾 Output File: {output_file}")
        print(f"   Status: Not created yet")
        print(f"   Processed: 0 rows")
        print(f"   Remaining: {total_rows} rows")
        print(f"\n📊 Progress: 0%")
        return
    
    wb_output = load_workbook(output_file, read_only=True)
    ws_output = wb_output.active
    processed_rows = ws_output.max_row - 1  # Exclude header
    wb_output.close()
    
    remaining = total_rows - processed_rows
    progress_pct = (processed_rows / total_rows * 100) if total_rows > 0 else 0
    
    print(f"\n💾 Output File: {output_file}")
    print(f"   Processed: {processed_rows} rows")
    print(f"   Remaining: {remaining} rows")
    
    print(f"\n📊 Progress: {progress_pct:.1f}%")
    
    # Progress bar
    bar_length = 40
    filled = int(bar_length * progress_pct / 100)
    bar = "█" * filled + "░" * (bar_length - filled)
    print(f"   [{bar}] {processed_rows}/{total_rows}")
    
    # Estimate
    if processed_rows > 0:
        avg_time_per_row = 3  # seconds (conservative estimate)
        remaining_time = remaining * avg_time_per_row
        hours = remaining_time // 3600
        minutes = (remaining_time % 3600) // 60
        
        print(f"\n⏱️  Estimated time remaining:")
        if hours > 0:
            print(f"   ~{int(hours)}h {int(minutes)}m")
        else:
            print(f"   ~{int(minutes)}m")
    
    # Next steps
    print(f"\n{'='*60}")
    if remaining > 0:
        print("📋 To continue processing:")
        print(f"   python resume_processing.py {source_file} 50")
        print("\n   Or process specific amount:")
        print(f"   python resume_processing.py {source_file} {min(100, remaining)}")
    else:
        print("✅ All rows processed!")
    
    print("="*60)


if __name__ == "__main__":
    import sys
    
    source_file = sys.argv[1] if len(sys.argv) > 1 else 'Transcript-24-11-2025.xlsx'
    output_file = sys.argv[2] if len(sys.argv) > 2 else 'IntentOfthetranscribetext.xlsx'
    
    check_progress(source_file, output_file)
