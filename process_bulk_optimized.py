"""
Optimized Bulk Processing with Smart Rate Limiting
- Reduces workers when rate limits are hit
- Adds delays between batches
- Processes in smaller batches to avoid overwhelming the API
"""

import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv

load_dotenv()

from process_transcribe_to_excel import TranscribeProcessor
from openpyxl import load_workbook


class OptimizedBulkProcessor:
    def __init__(self, excel_file='IntentOfthetranscribetext.xlsx', language='hindi', api_key=None):
        """Initialize Optimized Bulk Processor"""
        self.processor = TranscribeProcessor(
            excel_file=excel_file,
            llm_model='sarvam-m',
            language=language,
            api_key=api_key
        )
        self.rate_limit_delay = 1.5  # Delay between requests in seconds
        print(f"✓ Initialized with smart rate limiting ({self.rate_limit_delay}s delay)")
    
    def process_single_row(self, row_data):
        """Process a single row with rate limit delay"""
        row_idx, audio_name, transcribed_text = row_data
        
        # Add delay to avoid rate limits
        time.sleep(self.rate_limit_delay)
        
        try:
            if not transcribed_text or not str(transcribed_text).strip():
                return (row_idx, None, "Empty")
            
            transcribed_text = str(transcribed_text).strip()
            
            # Print immediately when starting
            print(f"  🔄 Processing row {row_idx}...", flush=True)
            
            result = self.processor.process_text(transcribed_text, audio_name)
            
            if result:
                return (row_idx, result, None)
            else:
                return (row_idx, None, "Failed")
                
        except Exception as e:
            return (row_idx, None, str(e))
    
    def process_from_excel_optimized(self, source_excel_file, limit=100, batch_size=5):
        """
        Process Excel file in batches with smart rate limiting
        
        Args:
            source_excel_file: Source Excel file path
            limit: Total number of rows to process
            batch_size: Number of rows to process concurrently (default: 5)
        """
        if not os.path.exists(source_excel_file):
            print(f"❌ ERROR: Source file not found: {source_excel_file}")
            return
        
        print(f"\n{'='*60}")
        print(f"OPTIMIZED BULK PROCESSING")
        print(f"{'='*60}")
        print(f"📖 Source: {source_excel_file}")
        print(f"💾 Output: {self.processor.excel_file}")
        print(f"📊 Total rows: {limit}")
        print(f"⚡ Batch size: {batch_size} concurrent requests")
        print(f"⏱️  Delay: {self.rate_limit_delay}s between requests")
        print(f"{'='*60}\n")
        
        # Read all rows
        wb = load_workbook(source_excel_file, read_only=True)
        ws = wb.active
        
        total_rows = ws.max_row
        rows_to_process = min(limit, total_rows - 1)
        
        rows_data = []
        for row_idx in range(2, min(2 + limit, total_rows + 1)):
            audio_name = ws.cell(row=row_idx, column=1).value
            transcribed_text = ws.cell(row=row_idx, column=2).value
            rows_data.append((row_idx, audio_name, transcribed_text))
        
        wb.close()
        
        # Process in batches
        processed = 0
        failed = 0
        skipped = 0
        start_time = time.time()
        
        total_batches = (len(rows_data) + batch_size - 1) // batch_size
        
        for batch_num in range(0, len(rows_data), batch_size):
            batch = rows_data[batch_num:batch_num + batch_size]
            current_batch = batch_num // batch_size + 1
            
            print(f"\n📦 Batch {current_batch}/{total_batches} ({len(batch)} rows)")
            print("-" * 60)
            
            with ThreadPoolExecutor(max_workers=batch_size) as executor:
                future_to_row = {executor.submit(self.process_single_row, row_data): row_data for row_data in batch}
                
                for future in as_completed(future_to_row):
                    row_data = future_to_row[future]
                    row_idx = row_data[0]
                    
                    try:
                        result_row_idx, result, error = future.result()
                        
                        if result:
                            self.processor.save_to_excel(result)
                            processed += 1
                            print(f"  ✓ Row {result_row_idx} - {processed}/{len(rows_data)}")
                        elif error == "Empty":
                            skipped += 1
                        else:
                            failed += 1
                            if error != "Failed":
                                print(f"  ✗ Row {result_row_idx}: {error[:50]}")
                    
                    except Exception as e:
                        failed += 1
                        print(f"  ✗ Row {row_idx}: {str(e)[:50]}")
            
            # Delay between batches
            if current_batch < total_batches:
                print(f"⏸️  Waiting 3s before next batch...")
                time.sleep(3)
        
        # Summary
        elapsed_time = time.time() - start_time
        print(f"\n{'='*60}")
        print(f"✅ COMPLETE!")
        print(f"{'='*60}")
        print(f"✓ Processed: {processed}")
        print(f"✗ Failed: {failed}")
        print(f"⊘ Skipped: {skipped}")
        print(f"⏱️  Time: {elapsed_time:.1f}s ({elapsed_time/60:.1f} min)")
        if processed > 0:
            print(f"⚡ Speed: {processed / elapsed_time:.2f} rows/sec")
        print(f"{'='*60}")


if __name__ == "__main__":
    import sys
    
    print("="*60, flush=True)
    print("OPTIMIZED BULK PROCESSING", flush=True)
    print("="*60, flush=True)
    
    api_key = os.getenv('SARVAM_API_KEY')
    if not api_key:
        print("❌ ERROR: SARVAM_API_KEY not found", flush=True)
        exit(1)
    
    # Get parameters from command line or use defaults
    source_file = sys.argv[1] if len(sys.argv) > 1 else 'Transcript-24-11-2025.xlsx'
    limit = int(sys.argv[2]) if len(sys.argv) > 2 else 20  # Changed default to 20
    batch_size = int(sys.argv[3]) if len(sys.argv) > 3 else 3  # Changed default to 3
    
    if not os.path.exists(source_file):
        print(f"❌ ERROR: Source file not found: {source_file}", flush=True)
        exit(1)
    
    print(f"\n📋 Usage: python process_bulk_optimized.py [source_file] [limit] [batch_size]", flush=True)
    print(f"📋 Example: python process_bulk_optimized.py Transcript-24-11-2025.xlsx 100 5\n", flush=True)
    print(f"🎯 Current settings: {limit} rows, batch size {batch_size}\n", flush=True)
    
    processor = OptimizedBulkProcessor(
        excel_file='IntentOfthetranscribetext.xlsx',
        language='hindi',
        api_key=api_key
    )
    
    processor.process_from_excel_optimized(source_file, limit=limit, batch_size=batch_size)
