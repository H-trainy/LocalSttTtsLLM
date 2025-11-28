"""
Resume Processing Helper
Automatically detects where you left off and continues from there
"""

import os
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

from process_simple import process_excel_simple


def find_last_processed_row(output_file='IntentOfthetranscribetext.xlsx'):
    """Find the last row that was processed"""
    if not os.path.exists(output_file):
        print(f"Output file not found. Starting from beginning.")
        return 2  # Start from row 2 (after header)
    
    try:
        wb = load_workbook(output_file, read_only=True)
        ws = wb.active
        last_row = ws.max_row
        wb.close()
        
        # Next row to process is last_row + 1
        # But we need to account for the header row
        next_row = last_row + 1
        
        print(f"✓ Found {last_row - 1} processed rows in output file")
        print(f"📍 Will resume from row {next_row}")
        
        return next_row
        
    except Exception as e:
        print(f"Error reading output file: {e}")
        print("Starting from beginning.")
        return 2


def main():
    import sys
    
    print("="*60)
    print("RESUME PROCESSING")
    print("="*60)
    
    # Get parameters
    source_file = sys.argv[1] if len(sys.argv) > 1 else 'Transcript-24-11-2025.xlsx'
    limit = int(sys.argv[2]) if len(sys.argv) > 2 else 50  # How many MORE rows to process
    
    if not os.path.exists(source_file):
        print(f"❌ ERROR: Source file not found: {source_file}")
        return
    
    # Find where to resume
    start_row = find_last_processed_row()
    
    # Check if already complete
    wb = load_workbook(source_file, read_only=True)
    ws = wb.active
    total_rows = ws.max_row
    wb.close()
    
    if start_row > total_rows:
        print(f"\n✅ All rows already processed!")
        print(f"Total rows in source: {total_rows}")
        print(f"Last processed row: {start_row - 1}")
        return
    
    remaining = total_rows - start_row + 1
    to_process = min(limit, remaining)
    
    print(f"\n📊 Status:")
    print(f"  Total rows in source: {total_rows}")
    print(f"  Already processed: {start_row - 2}")
    print(f"  Remaining: {remaining}")
    print(f"  Will process now: {to_process}")
    print(f"  Starting from row: {start_row}")
    
    # Confirm
    print(f"\n{'='*60}")
    response = input("Continue? (y/n): ").strip().lower()
    if response != 'y':
        print("Cancelled.")
        return
    
    # Process
    process_excel_simple(source_file, to_process, start_row)


if __name__ == "__main__":
    import sys
    
    print("\n📋 Usage: python resume_processing.py [source_file] [limit]")
    print("📋 Example: python resume_processing.py Transcript-24-11-2025.xlsx 50")
    print("📋 This will process 50 MORE rows from where you left off\n")
    
    main()
