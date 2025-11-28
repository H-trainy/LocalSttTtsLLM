"""
Simple Sequential Processing with Progress
Processes one row at a time with clear progress output
Most reliable - no rate limit issues
"""

import os
import sys
import time
from dotenv import load_dotenv

load_dotenv()

from process_transcribe_to_excel import TranscribeProcessor
from openpyxl import load_workbook


def process_excel_simple(source_file, limit=20, start_row=2):
    """
    Simple sequential processing with clear progress
    
    Args:
        source_file: Source Excel file
        limit: Number of rows to process
        start_row: Starting row number (default: 2, skip header)
    """
    print("="*60, flush=True)
    print("SIMPLE SEQUENTIAL PROCESSING", flush=True)
    print("="*60, flush=True)
    
    # Get API key
    api_key = os.getenv('SARVAM_API_KEY')
    if not api_key:
        print("❌ ERROR: SARVAM_API_KEY not found", flush=True)
        return
    
    # Check source file
    if not os.path.exists(source_file):
        print(f"❌ ERROR: Source file not found: {source_file}", flush=True)
        return
    
    print(f"\n📖 Source: {source_file}", flush=True)
    print(f"💾 Output: IntentOfthetranscribetext.xlsx", flush=True)
    print(f"📊 Processing {limit} rows starting from row {start_row}", flush=True)
    print(f"⏱️  Delay: 2 seconds between rows", flush=True)
    print("="*60 + "\n", flush=True)
    
    # Initialize processor
    print("Initializing...", flush=True)
    processor = TranscribeProcessor(
        excel_file='IntentOfthetranscribetext.xlsx',
        llm_model='sarvam-m',
        language='hindi',
        api_key=api_key
    )
    
    # Read source file
    print(f"\nReading {source_file}...", flush=True)
    wb = load_workbook(source_file, read_only=True)
    ws = wb.active
    total_rows = ws.max_row
    print(f"Total rows in file: {total_rows}", flush=True)
    
    # Process rows
    processed = 0
    failed = 0
    skipped = 0
    start_time = time.time()
    
    end_row = min(start_row + limit, total_rows + 1)
    
    print(f"\nStarting processing...\n", flush=True)
    
    for row_idx in range(start_row, end_row):
        current = row_idx - start_row + 1
        total = end_row - start_row
        
        print(f"{'='*60}", flush=True)
        print(f"[{current}/{total}] Processing Row {row_idx}", flush=True)
        print(f"{'='*60}", flush=True)
        
        try:
            # Get data
            audio_name = ws.cell(row=row_idx, column=1).value
            transcribed_text = ws.cell(row=row_idx, column=2).value
            
            # Skip empty
            if not transcribed_text or not str(transcribed_text).strip():
                print(f"⊘ Skipped (empty transcription)\n", flush=True)
                skipped += 1
                continue
            
            transcribed_text = str(transcribed_text).strip()
            
            # Show preview
            preview = transcribed_text[:60] + "..." if len(transcribed_text) > 60 else transcribed_text
            print(f"Text: {preview}", flush=True)
            if audio_name:
                print(f"Audio: {audio_name}", flush=True)
            
            # Process
            print("Processing...", flush=True)
            result = processor.process_text(transcribed_text, audio_name)
            
            if result:
                # Save
                processor.save_to_excel(result)
                processed += 1
                print(f"✓ Success!", flush=True)
                print(f"  Summary: {result['summary'][:50]}...", flush=True)
                print(f"  Intent: {result['intent']}", flush=True)
            else:
                failed += 1
                print(f"✗ Failed to process", flush=True)
            
        except Exception as e:
            failed += 1
            print(f"✗ Error: {str(e)[:100]}", flush=True)
        
        # Progress summary
        elapsed = time.time() - start_time
        rate = processed / elapsed if elapsed > 0 else 0
        remaining = total - current
        eta = remaining / rate if rate > 0 else 0
        
        print(f"\n📊 Progress: {processed} processed, {failed} failed, {skipped} skipped", flush=True)
        print(f"⏱️  Time: {elapsed:.0f}s | Speed: {rate:.2f} rows/s | ETA: {eta:.0f}s", flush=True)
        
        # Delay between rows
        if current < total:
            print(f"⏸️  Waiting 2 seconds...\n", flush=True)
            time.sleep(2)
    
    wb.close()
    
    # Final summary
    elapsed_time = time.time() - start_time
    print(f"\n{'='*60}", flush=True)
    print(f"✅ COMPLETE!", flush=True)
    print(f"{'='*60}", flush=True)
    print(f"✓ Processed: {processed}", flush=True)
    print(f"✗ Failed: {failed}", flush=True)
    print(f"⊘ Skipped: {skipped}", flush=True)
    print(f"⏱️  Total time: {elapsed_time:.1f}s ({elapsed_time/60:.1f} min)", flush=True)
    if processed > 0:
        print(f"⚡ Average speed: {processed / elapsed_time:.2f} rows/sec", flush=True)
    print(f"{'='*60}", flush=True)


if __name__ == "__main__":
    # Get parameters
    source_file = sys.argv[1] if len(sys.argv) > 1 else 'Transcript-24-11-2025.xlsx'
    limit = int(sys.argv[2]) if len(sys.argv) > 2 else 10  # Default: 10 rows
    start_row = int(sys.argv[3]) if len(sys.argv) > 3 else 2  # Default: start at row 2
    
    print("\n📋 Usage: python process_simple.py [source_file] [limit] [start_row]", flush=True)
    print("📋 Example: python process_simple.py Transcript-24-11-2025.xlsx 50 2\n", flush=True)
    
    process_excel_simple(source_file, limit, start_row)
