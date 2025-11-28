"""
Process Transcribed Text to Excel
Takes transcribed text, processes through LLM, extracts keywords and intent,
and saves to Excel file with columns: transcribe, llm response, keywords, intent
"""

import os
import sys

# Try to load from .env file
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed, continue without it
except Exception:
    pass  # .env file not found or error loading, continue

from llm_module import LLMModule
from intent_analyzer import IntentAnalyzer
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


class TranscribeProcessor:
    def __init__(self, excel_file='IntentOfthetranscribetext.xlsx', llm_model='openhathi-hi', language='hindi', api_key=None):
        """
        Initialize Transcribe Processor
        
        Args:
            excel_file: Path to Excel file (will be created if doesn't exist)
            llm_model: Sarvam AI model name (default: 'openhathi-hi' for Hindi)
                - 'openhathi-hi' (Hindi)
                - 'openhathi-en' (English)
            language: Language for analysis ('hindi', 'english', 'urdu', 'telugu')
            api_key: Sarvam AI API key (if None, reads from SARVAM_API_KEY env var)
        """
        self.excel_file = excel_file
        self.language = language.lower()
        
        print("Initializing LLM module with Sarvam AI...")
        # Use Sarvam's multilingual model (supports Hindi, English, and other Indian languages)
        model = llm_model or 'sarvam-m'
        
        self.llm = LLMModule(model_name=model, api_key=api_key)
        
        print("Initializing Intent Analyzer...")
        self.intent_analyzer = IntentAnalyzer(llm_model=model, api_key=api_key)
        
        # System prompts for LLM summary generation
        self.system_prompts = {
            'hindi': 'आप एक बुद्धिमान विश्लेषक हैं। दिए गए पाठ का संक्षिप्त और सटीक सारांश प्रदान करें। सारांश अंग्रेजी में होना चाहिए और केवल मुख्य बिंदु को एक वाक्य में बताएं। उदाहरण: "complain for unavailability of current"',
            'english': 'You are an intelligent analyst. Provide a brief and accurate summary of the given text. The summary should be in English and describe the main point in one sentence. Example: "complain for unavailability of current"',
            'urdu': 'آپ ایک ذہین تجزیہ کار ہیں۔ دیے گئے متن کا مختصر اور درست خلاصہ فراہم کریں۔ خلاصہ انگریزی میں ہونا چاہیے اور صرف اہم نکتہ کو ایک جملے میں بیان کریں۔',
            'telugu': 'మీరు ఒక తెలివైన విశ్లేషకుడు. ఇచ్చిన టెక్స్ట్‌కు సంక్షిప్తమైన మరియు ఖచ్చితమైన సారాంశం అందించండి. సారాంశం ఇంగ్లీష్‌లో ఉండాలి మరియు ముఖ్యమైన అంశాన్ని ఒక వాక్యంలో వివరించండి.'
        }
        
        # Initialize Excel file
        self._init_excel_file()
        
        print("Transcribe Processor initialized successfully!")
    
    def _detect_language_from_text(self, text):
        """
        Detect language from text using Unicode character ranges
        
        Args:
            text: Text to analyze
        
        Returns:
            Detected language ('hindi', 'english', 'urdu', 'telugu')
        """
        if not text or not text.strip():
            return self.language  # Return default language
        
        text = str(text).strip()
        
        # Count characters for each script
        telugu_chars = sum(1 for c in text if '\u0C00' <= c <= '\u0C7F')
        hindi_chars = sum(1 for c in text if '\u0900' <= c <= '\u097F')
        urdu_chars = sum(1 for c in text if '\u0600' <= c <= '\u06FF')
        english_chars = sum(1 for c in text if c.isalpha() and ord(c) < 128)
        
        total_chars = len([c for c in text if c.isalnum()])
        
        if total_chars == 0:
            return self.language  # Return default if no alphanumeric chars
        
        # Calculate percentages
        telugu_ratio = telugu_chars / total_chars if total_chars > 0 else 0
        hindi_ratio = hindi_chars / total_chars if total_chars > 0 else 0
        urdu_ratio = urdu_chars / total_chars if total_chars > 0 else 0
        english_ratio = english_chars / total_chars if total_chars > 0 else 0
        
        # Find language with highest ratio
        ratios = {
            'telugu': telugu_ratio,
            'hindi': hindi_ratio,
            'urdu': urdu_ratio,
            'english': english_ratio
        }
        
        detected_lang = max(ratios, key=ratios.get)
        
        # Only return detected language if ratio is significant (>10%)
        if ratios[detected_lang] > 0.1:
            return detected_lang
        else:
            return self.language  # Return default if no clear language detected
    
    def _init_excel_file(self):
        """Initialize Excel file with headers if it doesn't exist or is empty"""
        if os.path.exists(self.excel_file):
            try:
                wb = load_workbook(self.excel_file)
                ws = wb.active
                
                # Check if headers exist
                if ws.max_row == 0 or ws.cell(1, 1).value is None:
                    self._create_headers(ws)
                    wb.save(self.excel_file)
                    print(f"Created headers in existing Excel file: {self.excel_file}")
                else:
                    print(f"Using existing Excel file: {self.excel_file}")
            except Exception as e:
                print(f"Error reading Excel file: {e}. Creating new file...")
                self._create_new_excel()
        else:
            self._create_new_excel()
    
    def _create_new_excel(self):
        """Create a new Excel file with headers"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Transcriptions"
        self._create_headers(ws)
        wb.save(self.excel_file)
        print(f"Created new Excel file: {self.excel_file}")
    
    def _create_headers(self, ws):
        """Create column headers with styling"""
        headers = ['Audio Name', 'Transcribe', 'Summary', 'Intent']
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Set column widths
        ws.column_dimensions['A'].width = 40  # Audio Name
        ws.column_dimensions['B'].width = 50  # Transcribe
        ws.column_dimensions['C'].width = 50  # Summary
        ws.column_dimensions['D'].width = 30  # Intent
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def process_text(self, transcribed_text, audio_name=None):
        """
        Process a single transcribed text
        
        Args:
            transcribed_text: The transcribed text to process
            audio_name: Name of the audio file (optional)
        
        Returns:
            Dictionary with audio_name, transcribe, summary, intent
        """
        if not transcribed_text or not transcribed_text.strip():
            return None
        
        # Auto-detect language from text
        detected_language = self._detect_language_from_text(transcribed_text)
        if detected_language != self.language:
            print(f"Language detected: {detected_language.upper()} (was using {self.language.upper()})")
            processing_language = detected_language
        else:
            processing_language = self.language
        
        text_preview = transcribed_text[:50] if len(transcribed_text) > 50 else transcribed_text
        print(f"\nProcessing: {text_preview}...")
        if audio_name:
            print(f"Audio Name: {audio_name}")
        print(f"Using language: {processing_language.upper()}")
        
        # Step 1: Get LLM summary (always in English as per requirement)
        print("Generating summary with LLM...")
        # Use English system prompt for summary (always in English)
        system_prompt = self.system_prompts.get('english', self.system_prompts['english'])
        
        # Create prompt for summary (always in English as per requirement)
        summary_prompt = f"Summarize the following text in English in one sentence describing the main point:\n\n{transcribed_text}\n\nExample format: 'complain for unavailability of current'"
        
        summary = self.llm.generate(
            prompt=summary_prompt,
            system_prompt=system_prompt,
            max_tokens=100,  # Shorter for summary
            temperature=0.3  # Lower temperature for more focused summary
        )
        summary = summary.strip()
        # Clean up summary - remove quotes if present
        summary = summary.strip('"\'')
        summary_preview = summary[:50] if len(summary) > 50 else summary
        print(f"Summary: {summary_preview}...")
        
        # Step 2: Analyze for intent only (keywords removed)
        print("Extracting intent...")
        analysis = self.intent_analyzer.analyze(transcribed_text, language=processing_language)
        
        intent = analysis.get('intent', '')
        # Ensure intent is clean and short
        if intent:
            # Remove any extra words beyond 3
            words = intent.split()
            if len(words) > 3:
                intent = ' '.join(words[:3])
            intent = intent.strip().lower()
        
        print(f"Intent: {intent if intent else 'N/A'}")
        
        return {
            'audio_name': audio_name or '',
            'transcribe': transcribed_text,
            'summary': summary,
            'intent': intent
        }
    
    def save_to_excel(self, result):
        """
        Save result to Excel file
        
        Args:
            result: Dictionary with audio_name, transcribe, summary, intent
        """
        if not result:
            print("No result to save")
            return
        
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Find next empty row
            next_row = ws.max_row + 1
            
            # Add data
            ws.cell(row=next_row, column=1, value=result.get('audio_name', ''))
            ws.cell(row=next_row, column=2, value=result['transcribe'])
            ws.cell(row=next_row, column=3, value=result['summary'])
            ws.cell(row=next_row, column=4, value=result['intent'])
            
            # Set text wrapping for better readability
            for col in range(1, 5):
                cell = ws.cell(row=next_row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Save file
            wb.save(self.excel_file)
            # Don't print success message for each row to reduce clutter
            
        except Exception as e:
            print(f"[ERROR] Failed to save to Excel: {e}")
            raise
    
    def process_from_file(self, text_file):
        """
        Process transcribed text from a text file (one text per line)
        
        Args:
            text_file: Path to text file containing transcribed texts (one per line)
        """
        if not os.path.exists(text_file):
            print(f"Error: File not found: {text_file}")
            return
        
        print(f"Reading transcribed texts from: {text_file}")
        
        with open(text_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        total = len(lines)
        print(f"Found {total} transcribed texts to process\n")
        
        for idx, line in enumerate(lines, 1):
            transcribed_text = line.strip()
            if not transcribed_text:
                continue
            
            print(f"\n[{idx}/{total}] Processing...")
            result = self.process_text(transcribed_text)
            
            if result:
                self.save_to_excel(result)
            
            print("-" * 60)
    
    def process_single(self, transcribed_text, audio_name=None):
        """
        Process a single transcribed text and save to Excel
        
        Args:
            transcribed_text: The transcribed text to process
            audio_name: Name of the audio file (optional)
        """
        result = self.process_text(transcribed_text, audio_name)
        if result:
            self.save_to_excel(result)
        return result
    
    def process_from_excel(self, source_excel_file, limit=20, audio_col=1, transcribe_col=2):
        """
        Read transcribed texts from Excel file and process them
        
        Args:
            source_excel_file: Path to source Excel file with transcriptions
            limit: Number of rows to process (default: 20)
            audio_col: Column number for audio name (default: 1)
            transcribe_col: Column number for transcription (default: 2)
        """
        if not os.path.exists(source_excel_file):
            print(f"Error: Source Excel file not found: {source_excel_file}")
            return
        
        print(f"Reading from Excel file: {source_excel_file}")
        
        try:
            wb = load_workbook(source_excel_file, read_only=False)
            ws = wb.active
            
            total_rows = ws.max_row
            print(f"Total rows in source file: {total_rows}")
            
            # Process first 'limit' rows (skip header row)
            rows_to_process = min(limit, total_rows - 1)
            print(f"Processing first {rows_to_process} transcriptions...\n")
            
            processed = 0
            skipped = 0
            
            for row_idx in range(2, min(2 + limit, total_rows + 1)):
                try:
                    # Get audio name and transcription
                    audio_name = ws.cell(row=row_idx, column=audio_col).value
                    transcribed_text = ws.cell(row=row_idx, column=transcribe_col).value
                    
                    # Skip if no transcription
                    if not transcribed_text or not str(transcribed_text).strip():
                        skipped += 1
                        print(f"[{row_idx-1}] Skipping row {row_idx} - No transcription")
                        continue
                    
                    transcribed_text = str(transcribed_text).strip()
                    
                    print(f"\n[{processed + 1}/{rows_to_process}] Processing row {row_idx}...")
                    if audio_name:
                        print(f"Audio Name: {audio_name}")
                    
                    # Process the text
                    result = self.process_text(transcribed_text, audio_name)
                    
                    if result:
                        self.save_to_excel(result)
                        processed += 1
                        print(f"✓ Row {row_idx} processed and saved")
                    else:
                        skipped += 1
                        print(f"✗ Row {row_idx} failed to process")
                    
                    print("-" * 60)
                    
                except Exception as e:
                    print(f"[ERROR] Failed to process row {row_idx}: {e}")
                    import traceback
                    traceback.print_exc()
                    skipped += 1
                    continue
            
            print(f"\n{'='*60}")
            print(f"Processing Complete!")
            print(f"Processed: {processed}")
            print(f"Skipped: {skipped}")
            print(f"Total: {processed + skipped}")
            print(f"{'='*60}")
            
        except Exception as e:
            print(f"[ERROR] Failed to read Excel file: {e}")
            raise
    
    def process_multiple(self, transcribed_texts):
        """
        Process multiple transcribed texts
        
        Args:
            transcribed_texts: List of transcribed texts
        """
        total = len(transcribed_texts)
        print(f"Processing {total} transcribed texts\n")
        
        for idx, text in enumerate(transcribed_texts, 1):
            print(f"\n[{idx}/{total}] Processing...")
            result = self.process_text(text)
            if result:
                self.save_to_excel(result)
            print("-" * 60)


def main():
    """Main function for command-line usage"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Process transcribed text and save to Excel')
    parser.add_argument('--excel', '-e', default='IntentOfthetranscribetext.xlsx',
                       help='Excel file path (default: IntentOfthetranscribetext.xlsx)')
    parser.add_argument('--model', '-m', default='sarvam-m',
                       help='Sarvam AI model name (default: sarvam-m - multilingual model supporting Hindi, English, etc.)')
    parser.add_argument('--language', '-l', default='hindi',
                       choices=['hindi', 'english', 'urdu', 'telugu'],
                       help='Language for processing (default: hindi)')
    parser.add_argument('--api-key', type=str, default=None,
                       help='Sarvam AI API key (or set SARVAM_API_KEY environment variable)')
    parser.add_argument('--file', '-f', type=str,
                       help='Text file containing transcribed texts (one per line)')
    parser.add_argument('--text', '-t', type=str,
                       help='Single transcribed text to process')
    parser.add_argument('--source-excel', '-s', type=str,
                       help='Source Excel file to read transcriptions from')
    parser.add_argument('--limit', type=int, default=20,
                       help='Number of rows to process from source Excel (default: 20)')
    
    args = parser.parse_args()
    
    # Initialize processor
    processor = TranscribeProcessor(
        excel_file=args.excel,
        llm_model=args.model,
        language=args.language,
        api_key=args.api_key
    )
    
    # Process based on input
    if args.source_excel:
        processor.process_from_excel(args.source_excel, limit=args.limit)
    elif args.file:
        processor.process_from_file(args.file)
    elif args.text:
        processor.process_single(args.text)
    else:
        # Interactive mode (default)
        print("\n" + "="*60)
        print("Transcribe Text Processor - Interactive Mode")
        print("="*60)
        print("Using: Sarvam AI")
        print("\nEnter transcribed texts one by one")
        print("For each text, you will get:")
        print("  - Summary (in English, one sentence)")
        print("  - Intent (2-3 words, e.g., 'power cut', 'complaint')")
        print("\nCommands:")
        print("  - Type 'quit' or 'exit' to stop")
        print("  - Type 'excel:<path>' to process from Excel file")
        print("="*60 + "\n")
        
        while True:
            try:
                user_input = input("\nEnter transcribed text (or 'quit' to exit): ").strip()
                
                if user_input.lower() in ['quit', 'exit', 'q']:
                    print("Goodbye!")
                    break
                
                if user_input.lower().startswith('excel:'):
                    excel_path = user_input[6:].strip()
                    limit = 20  # Default limit
                    processor.process_from_excel(excel_path, limit=limit)
                    continue
                
                if not user_input:
                    print("Please enter some text or 'quit' to exit")
                    continue
                
                # Process single text
                result = processor.process_single(user_input)
                
                if result:
                    print("\n" + "="*60)
                    print("RESULT:")
                    print("="*60)
                    print(f"Transcribe: {result['transcribe']}")
                    print(f"Summary: {result['summary']}")
                    print(f"Intent: {result['intent']}")
                    print("="*60)
                
            except KeyboardInterrupt:
                print("\n\nInterrupted by user. Goodbye!")
                break
            except Exception as e:
                print(f"\n[ERROR] {e}")


if __name__ == "__main__":
    main()

