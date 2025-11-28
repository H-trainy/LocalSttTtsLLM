"""
Fast Bulk Processing with Concurrent API Calls
Process multiple rows from Excel in parallel to speed up processing
"""

import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv

# Load .env file
load_dotenv()

from process_transcribe_to_excel import TranscribeProcessor
from openpyxl import load_workbook


class FastBulkProcessor:
    def __init__(self, excel_file='IntentOfthetranscribetext.xlsx', language='hindi', api_key=None, max_workers=3):
        """
        Initialize Fast Bulk Processor
        
        Args:
            excel_file: Output Excel file path
            language: Language for processing
            api_key: Sarvam AI API key
            max_workers: Number of concurrent workers (default: 3, don't set too high to avoid rate limits)
        """
        self.processor = TranscribeProcessor(
            excel_file=excel_file,
            llm_model='sarvam-m',
            language=language,
            api_key=api_key
        )
        self.max_workers = max_workers
        print(f"✓ Initialized with {max_workers} concurrent workers")
    
    def process_single_row(self, row_data):
        """Process a single row (used by thread pool)"""
        row_idx, audio_name, transcribed_text = row_data
        
        try:
            if not transcribed_text or not str(transcribed_text).strip():
                return (row_idx, None, "Empty transcription")
            
            transcribed_text = str(transcribed_text).strip()
            
            # Process the text
            result = self.processor.process_text(transcribed_text, audio_name)
            
            if result:
                return (row_idx, result, None)
            else:
                return (row_idx, None, "Processing failed")
                
        except Exception as e:
            return (row_idx, None, str(e))
    
    def process_from_excel_fast(self, source_excel_file, limit=20, audio_col=1, transcribe_col=2):
        """
        Process Excel file with concurrent API calls for faster processing
        
        Args:
            source_excel_file: Path to source Excel file
            limit: Number of rows to process
            audio_col: Column number for audio name
            transcribe_col: Column number for transcription
        """
        if not os.path.exists(source_excel_file):
            print(f"❌ ERROR: Source file not found: {source_excel_file}")
            return
        
        print(f"\n{'='*60}")
        print(f"FAST BULK PROCESSING MODE")
        print(f"{'='*60}")
        print(f"📖 Reading from: {source_excel_file}")
        print(f"💾 Saving to: {self.processor.excel_file}")
        print(f"⚡ Using {self.max_workers} concurrent workers")
        print(f"📊 Processing up to {limit} rows...")
        print(f"{'='*60}\n")
        
        # Read all rows first
        wb = load_workbook(source_excel_file, read_only=True)
        ws = wb.active
        
        total_rows = ws.max_row
        rows_to_process = min(limit, total_rows - 1)
        
        # Collect all rows to process
        rows_data = []
        for row_idx in range(2, min(2 + limit, total_rows + 1)):
            audio_name = ws.cell(row=row_idx, column=audio_col).value
            transcribed_text = ws.cell(row=row_idx, column=transcribe_col).value
            rows_data.append((row_idx, audio_name, transcribed_text))
        
        wb.close()
        
        print(f"Collected {len(rows_data)} rows to process\n")
        
        # Process rows concurrently
        processed = 0
        failed = 0
        skipped = 0
        start_time = time.time()
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Submit all tasks
            future_to_row = {executor.submit(self.process_single_row, row_data): row_data for row_data in rows_data}
            
            # Process completed tasks
            for future in as_completed(future_to_row):
                row_data = future_to_row[future]
                row_idx = row_data[0]
                
                try:
                    result_row_idx, result, error = future.result()
                    
                    if result:
                        # Save to Excel
                        self.processor.save_to_excel(result)
                        processed += 1
                        print(f"✓ [{processed + failed + skipped}/{len(rows_data)}] Row {result_row_idx} processed")
                    elif error:
                        failed += 1
                        if error != "Empty transcription":
                            print(f"✗ [{processed + failed + skipped}/{len(rows_data)}] Row {result_row_idx} failed: {error}")
                        else:
                            skipped += 1
                    
                except Exception as e:
                    failed += 1
                    print(f"✗ Row {row_idx} exception: {e}")
        
        # Summary
        elapsed_time = time.time() - start_time
        print(f"\n{'='*60}")
        print(f"PROCESSING COMPLETE!")
        print(f"{'='*60}")
        print(f"✓ Processed: {processed}")
        print(f"✗ Failed: {failed}")
        print(f"⊘ Skipped: {skipped}")
        print(f"⏱️  Time: {elapsed_time:.2f}s")
        print(f"⚡ Speed: {processed / elapsed_time:.2f} rows/second")
        print(f"{'='*60}")


def main():
    """Main function"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Fast bulk processing with concurrent API calls')
    parser.add_argument('--source', '-s', required=True, help='Source Excel file')
    parser.add_argument('--output', '-o', default='IntentOfthetranscribetext.xlsx', help='Output Excel file')
    parser.add_argument('--limit', '-l', type=int, default=20, help='Number of rows to process')
    parser.add_argument('--workers', '-w', type=int, default=3, help='Number of concurrent workers (default: 3)')
    parser.add_argument('--language', type=str, default='hindi', help='Language for processing')
    
    args = parser.parse_args()
    
    # Get API key
    api_key = os.getenv('SARVAM_API_KEY')
    if not api_key:
        print("❌ ERROR: SARVAM_API_KEY not found in .env file")
        return
    
    # Initialize processor
    processor = FastBulkProcessor(
        excel_file=args.output,
        language=args.language,
        api_key=api_key,
        max_workers=args.workers
    )
    
    # Process
    processor.process_from_excel_fast(args.source, limit=args.limit)


if __name__ == "__main__":
    # Quick test mode
    print("="*60)
    print("FAST BULK PROCESSING")
    print("="*60)
    
    api_key = os.getenv('SARVAM_API_KEY')
    if not api_key:
        print("❌ ERROR: SARVAM_API_KEY not found")
        exit(1)
    
    # Process with 3 concurrent workers (adjust based on your rate limits)
    processor = FastBulkProcessor(
        excel_file='IntentOfthetranscribetext.xlsx',
        language='hindi',
        api_key=api_key,
        max_workers=3  # Start with 3, increase if no rate limit errors
    )
    
    source_file = 'Transcript-24-11-2025.xlsx'
    if not os.path.exists(source_file):
        print(f"❌ ERROR: Source file not found: {source_file}")
        exit(1)
    
    # Process first 20 rows (change as needed)
    processor.process_from_excel_fast(source_file, limit=20)
